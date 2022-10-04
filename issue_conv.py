
import openpyxl
import csv

from CustomCombobox import ComboBox

from PyQt5.QtCore import QRect, QSize, QMetaObject, QCoreApplication
from PyQt5.QtWidgets import QFileDialog, QWidget, QGridLayout, QLineEdit, QApplication, QTextBrowser, QPushButton, \
                            QMenuBar, QMenu, QAction, QMainWindow, QComboBox

import sys
import os


class UiMainWindow(object):
    # main excel workbook fields
    workbook = openpyxl.Workbook()
    wb_sheet = workbook.active
    wb_current_row = 2
    wb_path = ''
    csv_path = ''
    #-------------------------------------------------

    articles = []

    association_table = ""

    buff_tome = ''
    buff_issue = ''
    buff_price = ''

    article_data = ('articleid', 'articlecat', 'articletitle', 'alias', 'associations', 'articlelang', 'articleintrotext',
                    'god', 'tom', 'N vypuska', 'skvoznoj-nomer-vypuska', 'doi-stati', 'tip-stati', 'stranitsy-stati', 'avtory-stati',
                    'Об авторах', 'klychevye-slova', 'bibliografiya', 'dlya-tsitirovaniya', 'Tags', 'state', 'price', "link", "associations_category_id")

    article_data_holder = {
        'articleid': '',
        'articlecat': '',
        'articletitle': '',
        'alias': '',
        'associations': '',
        'articlelang': '',
        'articleintrotext': '',
        'god': '',
        'tom': '',
        'N vypuska': '',
        'skvoznoj-nomer-vypuska': '',
        'doi-stati': '',
        'tip-stati': '',
        'stranitsy-stati': '',
        'avtory-stati': '',
        'Об авторах': '',
        'klychevye-slova': '',
        'bibliografiya': '',
        'dlya-tsitirovaniya': '',
        'Tags': '',
        'state': '',
        'price': '',
        'link': '',
        'associations_category_id': ''
    }

    language_tags = ('<rus>\n', '<eng>\n')

    opening_tags = ('<doi>\n', '<article_type>\n', '<title>\n', '<keywords>\n', '<author>\n', '<reference>\n',
                    '<citation>\n', '<annotation>\n', "<rubric>\n")

    closing_tags = ('</doi>\n', '</article_type>\n', '</title>\n', '</keywords>\n', '</author>\n', '</reference>\n',
                    '</citation>\n', '</annotation>\n', "</rubric>\n")

    def setup_ui(self, window_main):
        # window properties
        window_main.setObjectName("window_main")
        window_main.resize(420, 290)
        window_main.setMinimumSize(QSize(420, 290))
        window_main.setMaximumSize(QSize(420, 290))
        #-------------------------------------------------
        
        # central widget
        self.centralwidget = QWidget(window_main)
        self.centralwidget.setObjectName("centralwidget")
        #-------------------------------------------------

        window_main.setCentralWidget(self.centralwidget)

        # status layout
        self.gridLayoutWidget = QWidget(self.centralwidget)
        self.gridLayoutWidget.setGeometry(QRect(10, 10, 401, 141))
        self.gridLayoutWidget.setObjectName("gridLayoutWidget")
        
        self.status_layout = QGridLayout(self.gridLayoutWidget)
        self.status_layout.setContentsMargins(0, 0, 0, 0)
        self.status_layout.setObjectName("status_layout")
        #-------------------------------------------------

        # status window
        self.textStatus = QTextBrowser(self.gridLayoutWidget)
        self.textStatus.setObjectName("textStatus")
        
        self.status_layout.addWidget(self.textStatus, 0, 0, 1, 0)
        #-------------------------------------------------

        # additional layout
        self.gridLayoutWidget_2 = QWidget(self.centralwidget)
        self.gridLayoutWidget_2.setGeometry(QRect(10, 160, 160, 100))
        self.gridLayoutWidget_2.setObjectName("gridLayoutWidget_2")

        self.additional_layout = QGridLayout(self.gridLayoutWidget_2)
        self.additional_layout.setContentsMargins(0, 0, 0, 0)
        self.additional_layout.setObjectName("additional_layout")
        #-------------------------------------------------

        # associations fields
        self.comboboxAssociation = ComboBox(self.gridLayoutWidget)
        self.comboboxAssociation.setPlaceholderText("-- Выбрать журнал --")
        
        self.status_layout.addWidget(self.comboboxAssociation, 1, 0, 1, 1)
        #-------------------------------------------------

        # manually filled fields
        self.lineTome = QLineEdit(self.gridLayoutWidget_2)
        self.lineTome.setAccessibleDescription("")
        self.lineTome.setInputMask("")
        self.lineTome.setText("")
        self.lineTome.setObjectName("lineTome")
        
        self.additional_layout.addWidget(self.lineTome, 0, 0, 1, 1)
        
        self.lineIssue = QLineEdit(self.gridLayoutWidget_2)
        self.lineIssue.setText("")
        self.lineIssue.setObjectName("lineIssue")
        
        self.additional_layout.addWidget(self.lineIssue, 2, 0, 1, 1)

        self.linePrice = QLineEdit(self.gridLayoutWidget_2)
        self.linePrice.setAccessibleDescription("")
        self.linePrice.setInputMask("")
        self.linePrice.setText("")
        self.linePrice.setObjectName("linePrice")
        
        self.additional_layout.addWidget(self.linePrice, 3, 0, 1, 1)
        #-------------------------------------------------

        # manually filled fields submit button
        self.btnAdd = QPushButton(self.gridLayoutWidget_2)
        self.btnAdd.setObjectName("btnAdd")
        
        self.additional_layout.addWidget(self.btnAdd, 4, 0, 1, 1)
        #-------------------------------------------------

        # table generation button
        self.btnGenerate = QPushButton(self.centralwidget)
        self.btnGenerate.setEnabled(False)
        self.btnGenerate.setGeometry(QRect(190, 170, 221, 81))
        self.btnGenerate.setObjectName("btnGenerate")
        #-------------------------------------------------

        # top menu bar
        self.menubar = QMenuBar(window_main)
        self.menubar.setGeometry(QRect(0, 0, 420, 21))
        self.menubar.setObjectName("menubar")
        #-------------------------------------------------
        
        # file menu
        self.file = QMenu(self.menubar)
        self.file.setObjectName("file")
        #-------------------------------------------------
        
        window_main.setMenuBar(self.menubar)

        # file menu actions
        self.fileOpen = QAction()
        self.fileOpen.setObjectName("fileOpen")
        self.file.addAction(self.fileOpen)

        self.menubar.addMenu(self.file)
        #-------------------------------------------------

        self.re_translate_ui(window_main)
        QMetaObject.connectSlotsByName(window_main)

        self.load_functions()

# loading of button functions
    def load_functions(self):
        self.fileOpen.triggered.connect(self.load_files)
        self.btnAdd.clicked.connect(self.tome_issue_price_respond)
        self.btnGenerate.clicked.connect(self.generate_excel)
        self.comboboxAssociation.currentTextChanged.connect(self.check_validity)
        self.get_journals()

# journals
    def get_journals(self):
        try:
            files = os.listdir("Журналы")
        except Exception:
            os.mkdir("Журналы")

        files = os.listdir("Журналы")
        journal_paths = []
        journal_names = []

        for filename in files:
            if filename.endswith(".xlsx"):
                journal_names.append(filename.split(".")[0])
                journal_paths.append("Журналы\\" + filename)

        self.journal_dict = dict(zip(journal_names, journal_paths))

        self.comboboxAssociation.addItems(journal_names)

# generation of excel table base
    def generate_base_table(self):
        for i in range(len(self.article_data)):
            self.wb_sheet.cell(row=self.wb_current_row - 1, column=i + 1).value = self.article_data[i]

        self.textStatus.append("Сгенерирован базовый лист...\n")

# generation of excel table
    def generate_excel(self):
        self.generate_base_table()
        self.process_article_categories()

        for file in self.articles:
            with open(file, 'r', encoding='utf-8-sig') as article:
                article_lines = article.readlines()

            # Data addition process

            self.add_static_info()

            rus_index = article_lines.index(self.language_tags[0])
            eng_index = article_lines.index(self.language_tags[1])

            for language_tag_id in range(len(self.language_tags)):

                if language_tag_id == 0:
                    source = article_lines[rus_index:eng_index:]
                elif language_tag_id == 1:
                    # WARNING Doesn't work with more than 2 language tags!
                    source = article_lines[eng_index::]

                self.add_dynamic_info(source, language_tag_id)

                #print(self.article_data_holder)

                self.add_to_table()

                self.wb_current_row += 1

        self.save_file()
            
        self.reset()

# file saving process
    def save_file(self):
        self.get_save_path()

        with open(self.csv_path, "w", encoding="utf-8-sig", newline="") as csv_file:

            writer = csv.writer(csv_file, quoting=csv.QUOTE_ALL)
            writer.writerows(self.wb_sheet.values)

        self.workbook.save(self.wb_path)

        self.textStatus.append("Файл сохранен! ({path})".format(path=self.wb_path))
        self.textStatus.append("Файл сохранен! ({path})".format(path=self.csv_path))
        self.textStatus.append("==============================================\n")

# info addition to the excel table
    def add_to_table(self):
        column_id = 1

        for key in self.article_data_holder:
            if len(self.article_data_holder[key]) > 0:
                self.wb_sheet.cell(row=self.wb_current_row, column=column_id).value = self.article_data_holder[key]

            column_id += 1

        self.textStatus.append("Информация добавлена ({language})...\n".format(language=self.article_data_holder["articlelang"]))

# static info addition
    def add_static_info(self):
        self.article_data_holder["tom"] = self.buff_tome
        self.article_data_holder["skvoznoj-nomer-vypuska"] = self.buff_issue
        self.article_data_holder["state"] = "1"
        self.article_data_holder["price"] = self.buff_price

# dynamic info addition
    def add_dynamic_info(self,  source, language_tag_id):
        if language_tag_id == 0:
            self.article_data_holder['articlelang'] = 'ru-RU'
        elif language_tag_id == 1:
            self.article_data_holder['articlelang'] = 'en-GB'

        for tag_id in range(len(self.opening_tags)):
            section_op = source.index(self.opening_tags[tag_id]) + 1
            section_cl = source.index(self.closing_tags[tag_id])

            if tag_id == 0:
                self.set_typo_info(section_op, section_cl, source, "doi-stati")

                self.set_year()
                self.set_pages()
                self.set_issue_number()
                self.set_association(language_tag_id)
                self.set_alias(language_tag_id)
                self.set_tags(language_tag_id)

                continue

            elif tag_id == 1:
                self.set_typo_info(section_op, section_cl, source, "tip-stati")
                continue
            elif tag_id == 2:
                self.set_typo_info(section_op, section_cl, source, "articletitle")
                continue
            elif tag_id == 3:
                self.set_typo_info(section_op, section_cl, source, "klychevye-slova")
                continue
            elif tag_id == 4:
                self.set_authors(source, tag_id)
                continue
            elif tag_id == 5:
                self.set_reference(section_op, section_cl, source)
                continue
            elif tag_id == 6:
                self.set_typo_info(section_op, section_cl, source, "dlya-tsitirovaniya")
                continue
            elif tag_id == 7:
                self.set_typo_info(section_op, section_cl, source, "articleintrotext", "<br>")
                continue
            elif tag_id == 8:
                self.set_category_fields(section_op, section_cl, source, language_tag_id)
                continue

    def process_article_categories(self):
        journal_path = self.journal_dict[self.comboboxAssociation.currentText()]

        category_wb = openpyxl.load_workbook(journal_path)
        category_sh = category_wb.active

        self.categories = {}

        row = 1
        col_a = 1
        col_c = 3

        while True:
            if category_sh.cell(row=row, column=col_a).value == None:
                break
            else:
                self.categories[category_sh.cell(row=row, column=col_a).value.upper()] = category_sh.cell(row=row, column=col_c).value
                row += 1

        cat_names_list = list(self.categories.keys())
        half_length = int(len(cat_names_list) / 2)
        self.category_comparison = dict.fromkeys(cat_names_list)
        pos = 0

        for key in self.category_comparison:
            if pos < half_length - 1:
                self.category_comparison[key] = cat_names_list[pos + half_length]
            else:
                self.category_comparison[key] = cat_names_list[pos - half_length]
            pos += 1

    def set_category_fields(self, section_op, section_cl, source, language_tag_id):
        category_name = source[section_op:section_cl]
        if len(category_name) > 0:
            category_name = category_name[0].split("\n")[0]
            category_name = str(category_name).upper()
        
        try:
            self.article_data_holder["articlecat"] = str(self.categories[category_name])

            if language_tag_id == 0:
                self.article_data_holder["associations_category_id"] = "en-GB#" + str(self.categories[self.category_comparison[category_name]])
            elif language_tag_id == 1:
                self.article_data_holder["associations_category_id"] = "ru-RU#" + str(self.categories[self.category_comparison[category_name]])
        except Exception:
            self.article_data_holder["articlecat"] = ""

            if language_tag_id == 0:
                self.article_data_holder["associations_category_id"] = "Категории нет в таблице!"
            elif language_tag_id == 1:
                self.article_data_holder["associations_category_id"] = "Категории нет в таблице!"

        


    def set_reference(self, section_op, section_cl, source):
        reference_list = source[section_op:section_cl:]

        previous_id = 0
        addition = 1

        for line_id in range(1, len(reference_list)):
            previous = int(reference_list[previous_id].split(".")[0])

            try:
                current = int(reference_list[line_id].split(".")[0])

                if current - previous == 1:
                    reference_list[line_id - 1] = reference_list[line_id - 1].replace("\n", '<br>')
                    previous_id += addition
                    addition = 1
                else:
                    reference_list[line_id - 1] = reference_list[line_id - 1].replace("\n", '')
                    addition += 1
            except ValueError:
                reference_list[line_id - 1] = reference_list[line_id - 1].replace("\n", '')
                addition += 1

        self.article_data_holder["bibliografiya"] = "".join(reference_list)

    def set_authors(self, source, tag_id):
        self.article_data_holder["avtory-stati"] = ""
        self.article_data_holder["Об авторах"] = ""

        next_author_exists = True

        while self.opening_tags[tag_id] in source:
            author_op = source.index(self.opening_tags[tag_id]) + 1
            author_cl = source.index(self.closing_tags[tag_id])

            author_source = " ".join(source[author_op:author_cl:]).replace("\n", "")
            author_name = author_source.split("–")[0][0:-1:]

            if author_name.count(".") < 2:
                author_words = author_name.split()
                author_name =  author_words[2] + " " + author_words[0][0] + ". " + author_words[1][0] + ". "

            self.article_data_holder["Об авторах"] += author_source
            self.article_data_holder["avtory-stati"] += author_name

            if self.opening_tags[tag_id] in source[author_cl + 1::]:
                source = source[author_cl + 1::]
            else:
                source = source[author_cl + 1::]
                next_author_exists = False

            if next_author_exists:
                self.article_data_holder["avtory-stati"] += ", "
                self.article_data_holder["Об авторах"] += "<br>"

    def set_issue_number(self):
        issue_number_string = str(int(self.article_data_holder["doi-stati"].split(".")[3]))
        self.article_data_holder["N vypuska"] = issue_number_string

    def set_pages(self):
        pages_string = self.article_data_holder["doi-stati"].split(".")[5]
        pages = pages_string.split("-")

        for page_id in range(len(pages)):
            page_symbols = list(pages[page_id])

            for page_symbol_id in range(len(page_symbols)):
                if page_symbols[page_symbol_id] == "0":
                    page_symbols[page_symbol_id] = ""
                else:
                    break

            pages[page_id] = "".join(page_symbols)

        self.article_data_holder["stranitsy-stati"] = "-".join(pages)

    def set_association(self, language_tag_id):
        association_string = self.article_data_holder['doi-stati'].split('/')[1].replace('.', '-')

        if language_tag_id == 0:
            self.article_data_holder['associations'] = 'en-GB#' + 'eng-' + association_string
        elif language_tag_id == 1:
            self.article_data_holder['associations'] = 'ru-RU#' + 'rus-' + association_string

    def set_year(self):
        #print(self.article_data_holder["doi-stati"])
        year = self.article_data_holder["doi-stati"].split(".")[2]
        self.article_data_holder["god"] = year

    def set_tags(self, language_tag_id):
        tag_string = "-".join([self.article_data_holder["tom"], self.article_data_holder["N vypuska"], self.article_data_holder["god"]])

        if language_tag_id == 0:
            self.article_data_holder['Tags'] = 'tom-' + tag_string
        elif language_tag_id == 1:
            self.article_data_holder['Tags'] = 'vol-' + tag_string

    def set_alias(self, language_tag_id):
        alias_string = self.article_data_holder["doi-stati"].split("/")[1].replace(".", "-")

        if language_tag_id == 0:
            self.article_data_holder['alias'] = 'rus-' + alias_string
        elif language_tag_id == 1:
            self.article_data_holder['alias'] = 'eng-' + alias_string

    def set_typo_info(self, section_op, section_cl, source, slot, swap=''):
        self.article_data_holder[slot] = ''

        for line_num in range(section_op, section_cl):
            line = source[line_num].replace('\n', swap)
            self.article_data_holder[slot] += line

# saving path creator
    def get_save_path(self):
        name_wb = self.article_data_holder['doi-stati'].split('/')[1].split('.pp.')[0] + ".xlsx"
        name_csv = self.article_data_holder['doi-stati'].split('/')[1].split('.pp.')[0] + ".csv"
        path = "/".join(self.articles[0].split('/')[:-1:])

        self.wb_path = path + "/" + name_wb
        self.csv_path = path + "/" + name_csv

# tome and issue validator
    def tome_issue_price_respond(self):
        self.buff_issue = ''
        self.buff_tome = ''
        self.buff_price = ''

        if len(self.lineIssue.text()) > 0 and len(self.lineTome.text()) > 0 and len(self.linePrice.text()) > 0:
            self.textStatus.append("Том, сквозной номер и цена успешно добавлены!")
            self.textStatus.append("==============================================\n")
            self.buff_issue = str(self.lineIssue.text())
            self.buff_tome = str(self.lineTome.text())
            self.buff_price = str(self.linePrice.text())
            self.lineIssue.clear()
            self.lineTome.clear()
            self.linePrice.clear()
        else:
            self.textStatus.append("Не введен том, сквозной номер или цена!")
            self.textStatus.append("==============================================\n")

        self.check_validity()

# general validator
    def check_validity(self):
        if len(self.articles) > 0 and (len(self.buff_tome) > 0 and len(self.buff_issue) > 0 and len(self.comboboxAssociation.currentText()) > 0):
            self.btnGenerate.setEnabled(True)

# loading association table
    # def load_association_table(self):
    #     self.association_table = ""
        
    #     file_dialog = QFileDialog()
    #     file_dialog.setFileMode(QFileDialog.ExistingFile)

    #     file_dialog.exec_()

    #     if len(file_dialog.selectedFiles()) > 0:
    #         self.association_table = file_dialog.selectedFiles()[0]
    #         self.lineAssociations.setText("<span style=\" color: #008000;\">" + self.association_table.split("/")[-1] + "</span>")
    #     else:
    #         self.association_table = ""

# loading files
    def load_files(self):
        self.articles.clear()
        file_dialog = QFileDialog()
        file_dialog.setFileMode(QFileDialog.ExistingFiles)

        file_dialog.exec_()

        self.articles = file_dialog.selectedFiles()

        if len(self.articles) > 0:
            self.textStatus.append("Открытые файлы:\n" + "\n".join(self.articles) + "\n")

        self.check_validity()

# resetment of data holders (preparation for next file processing)
# TODO: Optimize it, tbh ftm it works ass
    def reset(self):
        for key in self.article_data_holder:
            self.article_data_holder[key] = ''

        self.articles.clear()
        self.btnGenerate.setEnabled(False)
        
        self.workbook.remove(self.wb_sheet)
        self.workbook.create_sheet()
        self.wb_sheet = self.workbook.active

        self.wb_path = ''
        self.wb_current_row = 2
        self.buff_issue = ''
        self.buff_tome = ''


    def re_translate_ui(self, window_main):
        _translate = QCoreApplication.translate
        window_main.setWindowTitle(_translate("window_main", "Конвертер"))
        self.lineIssue.setPlaceholderText(_translate("window_main", "Сквозной номер"))
        self.lineTome.setPlaceholderText(_translate("window_main", "Том"))
        self.linePrice.setPlaceholderText(_translate("window_main", "Цена"))
        self.btnAdd.setText(_translate("window_main", "Добавить"))
        self.btnGenerate.setText(_translate("window_main", "Сгенерировать таблицу"))
        self.file.setTitle(_translate("window_main", "Файл"))
        self.fileOpen.setText(_translate("window_main", "Открыть файл для конвертирования"))
        self.fileOpen.setShortcut(_translate("window_main", "Ctrl+O"))
        '''
        self.fileOpenAssociationTable.setText(_translate("window_main", "Добавить таблицу соответствий"))
        self.fileOpenAssociationTable.setShortcut(_translate("window_main", "Ctrl+T"))
        '''


if __name__ == "__main__":
    app = QApplication(sys.argv)
    main_window = QMainWindow()
    ui = UiMainWindow()
    ui.setup_ui(main_window)
    main_window.show()
    sys.exit(app.exec_())
